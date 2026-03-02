"""
web2sheet — URL to Spreadsheet
Fetch strategy: ScraperAPI (primary) → Direct (fallback) → AI Extract (last resort)
"""

import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
import re, time, json, io, random
from urllib.parse import urljoin, urlparse
from anthropic import Anthropic

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="web2sheet",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS + Visual Effects ────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Syne:wght@600;700;800&family=DM+Mono:wght@300;400;500&display=swap');

/* ── TOKENS ── */
:root{
  --bg:#06040f; --bg2:#0d0818; --bg3:#110d1e;
  --v:#8b5cf6; --vl:#a78bfa; --vxl:#c4b5fd;
  --pk:#ec4899; --pkl:#f472b6;
  --cy:#22d3ee; --wh:#f0ebff; --mu:#6d5f8a;
  --br:rgba(139,92,246,0.18); --brb:rgba(139,92,246,0.45);
  --gv:rgba(139,92,246,0.35); --gpk:rgba(236,72,153,0.3);
}

/* ── BASE ── */
html,body,[class*="css"]{
  font-family:'DM Mono',monospace !important;
  background:var(--bg) !important; color:var(--wh) !important;
}
#MainMenu,footer,header{visibility:hidden;}
.block-container{padding-top:.5rem !important; max-width:1280px;}
h1,h2,h3{font-family:'Syne',sans-serif !important;}

/* ── CRT SCANLINES — fixed overlay ── */
body::before{
  content:''; position:fixed; inset:0; z-index:9990; pointer-events:none;
  background:repeating-linear-gradient(
    0deg, transparent, transparent 3px,
    rgba(0,0,0,0.07) 3px, rgba(0,0,0,0.07) 4px
  );
  animation:scanroll 14s linear infinite;
}
@keyframes scanroll{to{background-position:0 100vh;}}

/* ── VIGNETTE — fixed overlay ── */
body::after{
  content:''; position:fixed; inset:0; z-index:9989; pointer-events:none;
  background:radial-gradient(ellipse 82% 82% at 50% 50%,
    transparent 52%, rgba(6,4,15,0.86) 100%);
}

/* ── PARTICLE CANVAS ── */
#w2s-canvas{
  position:fixed; inset:0; z-index:0; pointer-events:none;
}

/* ── SIDEBAR ── */
section[data-testid="stSidebar"]{
  background:rgba(8,5,18,0.96) !important;
  border-right:1px solid var(--br) !important;
  backdrop-filter:blur(16px);
}
section[data-testid="stSidebar"] *{color:var(--wh) !important;}

/* ── NAV BAR (injected HTML element) ── */
.w2s-nav{
  position:sticky; top:0; z-index:9980;
  display:flex; align-items:center; justify-content:space-between;
  padding:0 24px; height:56px; margin-bottom:1.2rem;
  background:rgba(6,4,15,0.88); backdrop-filter:blur(20px);
  border-bottom:1px solid var(--br);
}
.w2s-nav-logo{
  font-family:'Space Mono',monospace; font-weight:700; font-size:.95rem;
  color:var(--vxl); text-shadow:0 0 22px var(--gv); letter-spacing:.5px;
}
.w2s-nav-logo .sl{color:var(--pk);}
.w2s-nav-links{display:flex; gap:28px; list-style:none; margin:0; padding:0;}
.w2s-nav-links a{
  font-family:'DM Mono',monospace; font-size:.6rem; letter-spacing:2.5px;
  text-transform:uppercase; color:var(--mu); text-decoration:none;
  position:relative; padding-bottom:2px; transition:color .2s;
}
.w2s-nav-links a::after{
  content:''; position:absolute; bottom:-1px; left:0; right:0; height:1px;
  background:linear-gradient(90deg,var(--v),var(--pk));
  transform:scaleX(0); transform-origin:left; transition:transform .3s;
}
.w2s-nav-links a:hover{color:var(--vxl);}
.w2s-nav-links a:hover::after{transform:scaleX(1);}
.w2s-status{
  font-family:'DM Mono',monospace; font-size:.58rem; letter-spacing:2px;
  color:var(--mu); display:flex; align-items:center; gap:7px;
}
.w2s-sdot{
  width:7px; height:7px; border-radius:50%;
  background:var(--v); display:inline-block;
  box-shadow:0 0 10px var(--v),0 0 20px var(--gv);
  animation:sdot 2.5s ease-in-out infinite;
}
@keyframes sdot{
  0%,100%{box-shadow:0 0 7px var(--v),0 0 16px var(--gv);}
  50%{box-shadow:0 0 18px var(--vl),0 0 32px var(--gv); opacity:.6;}
}

/* ── LOGO / HERO ── */
.logo{
  font-family:'Space Mono',monospace !important; font-size:2rem; font-weight:700;
  background:linear-gradient(135deg,var(--vxl),var(--pkl));
  -webkit-background-clip:text; -webkit-text-fill-color:transparent;
  background-clip:text; letter-spacing:1px;
  text-shadow:none;
}

/* GLITCH on logo */
.logo-wrap{position:relative; display:inline-block;}
.logo-wrap::before,.logo-wrap::after{
  content:attr(data-t); position:absolute; top:0; left:0;
  font-family:'Space Mono',monospace; font-size:2rem; font-weight:700;
  letter-spacing:1px; white-space:nowrap;
  background:none; -webkit-text-fill-color:unset;
}
.logo-wrap::before{color:var(--pk); animation:ga 6s infinite 2s; opacity:.65;}
.logo-wrap::after {color:var(--cy); animation:gb 6s infinite 2.05s; opacity:.45;}
@keyframes ga{
  0%,88%,100%{clip-path:inset(100% 0 0 0);transform:translate(0);}
  89%{clip-path:inset(20% 0 60% 0);transform:translate(-4px,2px);}
  91%{clip-path:inset(65% 0 15% 0);transform:translate(4px,-2px);}
  93%{clip-path:inset(35% 0 45% 0);transform:translate(-3px,1px);}
  95%{clip-path:inset(75% 0 8% 0);transform:translate(3px,-1px);}
}
@keyframes gb{
  0%,88%,100%{clip-path:inset(100% 0 0 0);transform:translate(0);}
  89%{clip-path:inset(55% 0 25% 0);transform:translate(3px,-2px);}
  91%{clip-path:inset(10% 0 78% 0);transform:translate(-4px,1px);}
  93%{clip-path:inset(45% 0 35% 0);transform:translate(2px,-1px);}
  95%{clip-path:inset(5% 0 85% 0);transform:translate(-3px,2px);}
}

/* ── DIVIDER ── */
.divider{border:none;border-top:1px solid var(--br);margin:1rem 0;}

/* ── GLASS CARDS ── */
.glass{
  background:rgba(13,8,24,0.82); border:1px solid var(--br);
  border-radius:10px; padding:1.2rem; margin-bottom:.8rem;
}

/* ── PRODUCT CARDS ── */
.pcard{
  background:rgba(13,8,24,0.9); border:1px solid var(--br);
  border-radius:10px; overflow:hidden; margin-bottom:.8rem;
  transition:border-color .25s, box-shadow .25s;
}
.pcard:hover{border-color:var(--brb); box-shadow:0 0 20px var(--gv);}
.pbody{padding:.85rem;}
.ptitle{font-family:'Syne',sans-serif;font-weight:700;font-size:.88rem;margin-bottom:.3rem;color:var(--wh);}
.pprice{color:#28c840;font-weight:700;font-size:.95rem;margin-bottom:.22rem;}
.prating{color:#febc2e;font-size:.78rem;margin-bottom:.22rem;}
.pdesc{color:var(--mu);font-size:.73rem;line-height:1.55;margin-bottom:.3rem;}

/* ── PILLS ── */
.pill{
  display:inline-block; padding:2px 8px;
  background:rgba(139,92,246,0.1); border:1px solid var(--br);
  border-radius:20px; font-size:.64rem; color:var(--vxl); margin:2px;
}

/* ── DETAIL TABLE ── */
.dtable{width:100%;border-collapse:collapse;font-size:.8rem;}
.dtable th{
  background:rgba(139,92,246,0.08); color:var(--vxl);
  padding:8px 12px; text-align:left;
  border-bottom:1px solid var(--br); font-weight:500;
}
.dtable td{
  padding:7px 12px; border-bottom:1px solid rgba(139,92,246,0.07);
  color:var(--mu); vertical-align:top;
}
.dtable tr:hover td{background:rgba(139,92,246,0.05); color:var(--wh);}
.dtable td:first-child{font-weight:600; color:var(--wh); width:28%;}

/* ── STAT BOXES ── */
.sbox{
  background:rgba(13,8,24,0.85); border:1px solid var(--br);
  border-radius:10px; padding:1rem; text-align:center;
  transition:border-color .25s;
}
.sbox:hover{border-color:var(--brb);}
.snum{
  font-family:'Syne',sans-serif; font-size:1.8rem; font-weight:800;
  background:linear-gradient(135deg,var(--vxl),var(--pkl));
  -webkit-background-clip:text; -webkit-text-fill-color:transparent;
  background-clip:text;
}
.slbl{font-size:.68rem; color:var(--mu); margin-top:2px; letter-spacing:1.5px; text-transform:uppercase;}

/* ── HEADING LEVEL HEADERS ── */
.cheader{
  display:flex; align-items:center; gap:10px; padding:.55rem 1rem;
  background:rgba(13,8,24,0.7); border:1px solid var(--br);
  border-radius:8px; margin-bottom:.6rem;
}
.ccount{font-size:.7rem; color:var(--mu); margin-left:auto;}

/* ── KEY BANNER ── */
.key-banner{
  background:rgba(139,92,246,0.07); border:1px solid rgba(139,92,246,0.25);
  border-radius:10px; padding:.9rem 1.1rem; margin-bottom:1rem;
}

/* ── AI BANNER ── */
.ai-banner{
  background:rgba(13,8,24,0.85); border:1px solid var(--brb);
  border-left:3px solid var(--v); border-radius:10px; padding:.9rem 1.1rem;
  margin-bottom:.8rem;
}

/* ── STRATEGY BADGE ── */
.strat-badge{
  display:inline-block; font-family:'DM Mono',monospace; font-size:.62rem;
  padding:3px 10px; border-radius:20px; letter-spacing:1.5px;
  border:1px solid; margin-bottom:.6rem;
}

/* ── STREAMLIT OVERRIDES ── */
.stButton > button{
  font-family:'DM Mono',monospace !important; letter-spacing:2px;
  text-transform:uppercase; font-size:.7rem !important;
  background:linear-gradient(135deg,var(--v),var(--pk)) !important;
  color:#fff !important; border:none !important; border-radius:4px !important;
  transition:all .25s !important;
}
.stButton > button:hover{
  filter:brightness(1.15) !important;
  box-shadow:0 0 24px var(--gv) !important;
}
.stTextInput > div > div > input{
  background:rgba(13,8,24,0.95) !important; border:1px solid var(--brb) !important;
  color:var(--wh) !important; font-family:'DM Mono',monospace !important;
  border-radius:4px !important;
}
.stTextInput > div > div > input:focus{border-color:var(--v) !important; box-shadow:0 0 12px var(--gv) !important;}
.stSelectbox > div,[data-baseweb="select"]{background:rgba(13,8,24,0.95) !important;}
.stTabs [data-baseweb="tab-list"]{
  background:rgba(13,8,24,0.8) !important; border-bottom:1px solid var(--br) !important;
  gap:2px;
}
.stTabs [data-baseweb="tab"]{
  font-family:'DM Mono',monospace !important; font-size:.68rem !important;
  letter-spacing:1.5px; color:var(--mu) !important; background:transparent !important;
  border-radius:0 !important; border-bottom:2px solid transparent !important;
  transition:all .2s;
}
.stTabs [aria-selected="true"]{
  color:var(--vxl) !important; border-bottom-color:var(--v) !important;
}
.stToggle label span{font-family:'DM Mono',monospace !important; font-size:.72rem !important;}
.stSlider .st-bd{background:var(--v) !important;}
div[data-testid="stStatusWidget"]{display:none;}
.stDataFrame{border:1px solid var(--br) !important; border-radius:8px !important;}
.stExpander{border:1px solid var(--br) !important; border-radius:8px !important; background:rgba(13,8,24,0.7) !important;}
.stAlert{border-radius:8px !important;}
</style>
""", unsafe_allow_html=True)

# ── Particle canvas + effects injected once ────────────────────────────────────
st.markdown("""
<canvas id="w2s-canvas"></canvas>

<div class="w2s-nav">
  <div class="w2s-nav-logo">web2sheet<span class="sl">/</span></div>
  <ul class="w2s-nav-links">
    <li><a href="#extract">Extract</a></li>
    <li><a href="#images">Images</a></li>
    <li><a href="#tables">Tables</a></li>
    <li><a href="#export">Export</a></li>
  </ul>
  <div class="w2s-status">
    <span class="w2s-sdot"></span>LIVE · ALL SYSTEMS
  </div>
</div>

<script>
(function(){
  if(document.getElementById('w2s-init')) return;
  var init = document.createElement('span');
  init.id = 'w2s-init';
  document.body.appendChild(init);

  var cvs = document.getElementById('w2s-canvas');
  if(!cvs) return;
  var ctx = cvs.getContext('2d');
  var W, H, pts=[], mx=-999, my=-999;

  function resize(){ W=cvs.width=window.innerWidth; H=cvs.height=window.innerHeight; }
  resize();
  window.addEventListener('resize', resize);

  var COLS=['#8b5cf6','#a78bfa','#ec4899','#f472b6','#c4b5fd','#7c3aed','#db2777'];

  function P(){
    this.reset = function(init){
      this.x = Math.random()*W;
      this.y = init ? Math.random()*H : H+8;
      this.r = Math.random()*2.2+0.5;
      this.vy = -(Math.random()*.45+.08);
      this.vx = (Math.random()-.5)*.16;
      this.a  = Math.random()*.5+.1;
      this.ph = Math.random()*Math.PI*2;
      this.c  = COLS[Math.floor(Math.random()*COLS.length)];
      this.sq = Math.random()>.72;
    };
    this.reset(true);
  }
  P.prototype.tick = function(){
    this.ph+=.02; this.y+=this.vy; this.x+=this.vx;
    var dx=this.x-mx, dy=this.y-my, d2=dx*dx+dy*dy;
    if(d2<10000){ var f=.5/Math.sqrt(d2+1); this.x+=dx*f; this.y+=dy*f; }
    if(this.y<-8)this.reset(false);
    if(this.x<0)this.x=W; if(this.x>W)this.x=0;
  };
  P.prototype.draw = function(){
    var a=this.a*(.65+.35*Math.sin(this.ph));
    ctx.save(); ctx.globalAlpha=a; ctx.fillStyle=this.c;
    if(this.sq) ctx.fillRect(this.x-this.r,this.y-this.r,this.r*2,this.r*2);
    else{ ctx.beginPath(); ctx.arc(this.x,this.y,this.r,0,Math.PI*2); ctx.fill(); }
    ctx.restore();
  };

  for(var i=0;i<120;i++) pts.push(new P());

  function drawWeb(){
    var MD=80;
    for(var i=0;i<pts.length;i++){
      for(var j=i+1;j<pts.length;j++){
        var dx=pts[i].x-pts[j].x, dy=pts[i].y-pts[j].y;
        var d=Math.sqrt(dx*dx+dy*dy);
        if(d<MD){
          ctx.save(); ctx.globalAlpha=.055*(1-d/MD);
          ctx.strokeStyle='#8b5cf6'; ctx.lineWidth=.4;
          ctx.beginPath(); ctx.moveTo(pts[i].x,pts[i].y);
          ctx.lineTo(pts[j].x,pts[j].y); ctx.stroke(); ctx.restore();
        }
      }
    }
  }

  function frame(){
    ctx.clearRect(0,0,W,H);
    drawWeb();
    pts.forEach(function(p){ p.tick(); p.draw(); });
    requestAnimationFrame(frame);
  }
  frame();

  document.addEventListener('mousemove', function(e){ mx=e.clientX; my=e.clientY; });
})();
</script>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
]

def _hdrs(ua=None):
    return {
        "User-Agent": ua or random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
    }

def _is_real_page(resp) -> bool:
    if resp.status_code != 200: return False
    t = resp.text
    if len(t) < 500: return False
    bad = ["cf-browser-verification","just a moment","enable javascript and cookies",
           "ddos protection","are you a robot","captcha","access denied","403 forbidden"]
    low = t.lower()
    return sum(1 for b in bad if b in low) < 2

# ─────────────────────────────────────────────────────────────────────────────
# FETCH STRATEGIES
# ─────────────────────────────────────────────────────────────────────────────

def fetch_via_scraperapi(url: str, key: str) -> str | None:
    """ScraperAPI — handles JS rendering, rotating proxies, anti-bot bypass."""
    try:
        endpoint = "https://api.scraperapi.com/"
        params = {"api_key": key, "url": url, "render": "false", "country_code": "in"}
        r = requests.get(endpoint, params=params, timeout=60)
        if r.status_code == 200 and len(r.text) > 500:
            return r.text
    except Exception:
        pass
    return None


def fetch_direct(url: str, delay=1.0) -> str | None:
    """Direct request with browser headers + homepage cookie priming."""
    time.sleep(delay)
    s = requests.Session()
    ua = random.choice(USER_AGENTS)
    try:
        parsed = urlparse(url)
        s.get(f"{parsed.scheme}://{parsed.netloc}", headers=_hdrs(ua), timeout=8)
        time.sleep(0.5)
    except Exception:
        pass
    for _ in range(2):
        try:
            r = s.get(url, headers=_hdrs(ua), timeout=20, allow_redirects=True)
            if _is_real_page(r):
                return r.text
        except Exception:
            pass
        time.sleep(1.5)
    return None


def fetch_google_cache(url: str) -> str | None:
    """Google's cached copy — bypasses many blocks."""
    try:
        cache = f"https://webcache.googleusercontent.com/search?q=cache:{url}&hl=en"
        r = requests.get(cache, headers=_hdrs(), timeout=20)
        if _is_real_page(r):
            return r.text
    except Exception:
        pass
    return None


def fetch_wayback(url: str) -> str | None:
    """Wayback Machine latest snapshot."""
    try:
        meta = requests.get(f"http://archive.org/wayback/available?url={url}", timeout=10).json()
        snap = meta.get("archived_snapshots",{}).get("closest",{}).get("url")
        if snap:
            r = requests.get(snap, headers=_hdrs(), timeout=25)
            if _is_real_page(r):
                return r.text
    except Exception:
        pass
    return None


def fetch_with_ai(url: str, client: Anthropic) -> tuple[str, str]:
    """
    Use Claude with web_search tool to extract structured data directly.
    Returns (fake_html_with_data, 'ai_search') — no HTTP fetch needed.
    """
    prompt = f"""You are a web data extractor. Search for the page at this URL and extract ALL product/item data you find.
URL: {url}

Return a clean HTML table with ALL items you find. Include these columns where available:
Title/Name, Price, Rating, Description, Image URL, Category, Brand, URL/Link

Format your entire response as a valid HTML <table> with <thead> and <tbody>.
Include as many rows as you can find. Do not add explanation, just the HTML table."""

    resp = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4000,
        tools=[{"type": "web_search_20250305", "name": "web_search"}],
        messages=[{"role": "user", "content": prompt}]
    )
    # Collect all text from response
    html_parts = []
    for block in resp.content:
        if block.type == "text":
            html_parts.append(block.text)
    full = "\n".join(html_parts)
    # Wrap in minimal HTML for BeautifulSoup
    return f"<html><body>{full}</body></html>", "ai_search"


def smart_fetch(url: str, scraperapi_key: str, use_ai_fallback: bool, client=None, delay=1.0) -> tuple[str, str]:
    """
    Master fetch: tries all strategies, returns (html, strategy_name).
    """
    # 1 — ScraperAPI (most reliable, works on 91mobiles, Amazon, etc.)
    if scraperapi_key:
        html = fetch_via_scraperapi(url, scraperapi_key)
        if html:
            return html, "scraperapi"

    # 2 — Direct
    html = fetch_direct(url, delay)
    if html:
        return html, "direct"

    # 3 — Google Cache
    html = fetch_google_cache(url)
    if html:
        return html, "google_cache"

    # 4 — Wayback Machine
    html = fetch_wayback(url)
    if html:
        return html, "wayback"

    # 5 — AI web search (no HTTP at all — Claude searches the web)
    if use_ai_fallback and client:
        try:
            html, strat = fetch_with_ai(url, client)
            return html, strat
        except Exception:
            pass

    raise Exception("All strategies failed. The site blocks all automated access including proxies.")


# ─────────────────────────────────────────────────────────────────────────────
# DATA EXTRACTION HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def resolve_img(src, base):
    if not src: return ""
    src = src.strip()
    if src.startswith("data:"): return ""
    if src.startswith("//"): return "https:" + src
    if src.startswith("http"): return src
    return urljoin(base, src)


def extract_images(soup, base):
    seen, imgs = set(), []
    for tag in soup.find_all("img"):
        src = (tag.get("src") or tag.get("data-src") or
               tag.get("data-lazy-src") or tag.get("data-original") or "")
        r = resolve_img(src, base)
        if r and r not in seen:
            ext = r.lower().split("?")[0]
            if any(ext.endswith(e) for e in [".jpg",".jpeg",".png",".webp",".gif",".svg"]):
                seen.add(r)
                imgs.append({"Image URL": r, "Alt Text": tag.get("alt",""),
                              "Width": tag.get("width",""), "Height": tag.get("height","")})
    return imgs


def extract_tables(soup):
    dfs = []
    for t in soup.find_all("table"):
        try:
            df = pd.read_html(str(t))[0]
            if not df.empty: dfs.append(df)
        except Exception: pass
    return dfs


def extract_meta(soup):
    m = {}
    m["Page Title"] = soup.title.get_text(strip=True) if soup.title else ""
    for tag in soup.find_all("meta"):
        n = tag.get("name") or tag.get("property") or ""
        c = tag.get("content") or ""
        if n and c: m[n] = c
    return m


def extract_links(soup, base):
    seen, rows = set(), []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        text = a.get_text(strip=True)
        full = urljoin(base, href)
        if full not in seen and text:
            seen.add(full)
            rows.append({"Link Text": text, "URL": full, "Domain": urlparse(full).netloc})
    return rows


def extract_headings(soup):
    rows = []
    for tag in ["h1","h2","h3","h4","h5","h6"]:
        for el in soup.find_all(tag):
            t = el.get_text(strip=True)
            if t: rows.append({"Level": tag.upper(), "Text": t})
    return rows


def extract_items(soup, base):
    price_re  = re.compile(r'[\$£€₹¥]\s?\d[\d,\.]*|\d[\d,\.]*\s*(?:USD|EUR|GBP|INR|rs\.?)', re.I)
    rating_re = re.compile(r'(\d[\.,]\d)\s*(?:out of|\/\d|stars?|★|rating)', re.I)

    selectors = ["article","[class*='product']","[class*='card']","[class*='item']",
                 "[class*='listing']","[class*='result']","[class*='post']","li"]
    candidates = []
    for sel in selectors:
        found = soup.select(sel)
        if len(found) >= 3:
            candidates = found[:150]; break

    items = []
    for el in candidates:
        text = el.get_text(separator=" ", strip=True)
        if len(text) < 15: continue
        row = {}
        for t in ["h1","h2","h3","h4","h5","a"]:
            tag = el.find(t)
            if tag: row["Title"] = tag.get_text(strip=True)[:220]; break
        if "Title" not in row: row["Title"] = text[:120]

        img = el.find("img")
        row["Image URL"] = resolve_img(
            img.get("src") or img.get("data-src") or img.get("data-lazy-src") or "" if img else "", base)
        row["Image Alt"] = img.get("alt","") if img else ""

        pm = price_re.search(text)
        row["Price"] = pm.group(0).strip() if pm else ""
        rm = rating_re.search(text)
        row["Rating"] = rm.group(1).replace(",",".") if rm else ""

        link = el.find("a", href=True)
        row["URL"] = urljoin(base, link["href"]) if link else ""

        paras = [p.get_text(strip=True) for p in el.find_all("p") if len(p.get_text(strip=True))>20]
        row["Description"] = paras[0][:300] if paras else ""

        row["Category"] = (el.get("data-category") or el.get("data-type") or
                           el.get("data-section") or "")
        row["Brand"]    = el.get("data-brand") or el.get("data-seller") or ""
        row["SKU"]      = el.get("data-sku") or el.get("data-id") or ""

        am = re.search(r'in stock|out of stock|available|sold out', text, re.I)
        row["Availability"] = am.group(0).title() if am else ""
        items.append(row)
    return items


def ai_analyze(html, client):
    prompt = f"""Analyze this HTML. Return ONLY valid JSON (no markdown):
{{"page_type":"e-commerce|news|directory|blog|jobs|real-estate|other",
"summary":"1 sentence",
"categories":["list"],
"key_fields":["fields found"]}}
HTML:{html[:3000]}"""
    r = client.messages.create(model="claude-sonnet-4-20250514", max_tokens=400,
        messages=[{"role":"user","content":prompt}])
    raw = re.sub(r"```json|```","",r.content[0].text.strip()).strip()
    try: return json.loads(raw)
    except: return {"page_type":"other","summary":"","categories":[],"key_fields":[]}


def clean_df(df, trim=True, dedup_col=None):
    if trim: df = df.apply(lambda c: c.map(lambda x: x.strip() if isinstance(x,str) else x))
    if dedup_col and dedup_col in df.columns:
        df = df.drop_duplicates(subset=[dedup_col])
    return df.reset_index(drop=True)


def to_excel(df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    buf = io.BytesIO()
    wb = Workbook(); ws = wb.active; ws.title = "web2sheet"
    hf = PatternFill("solid",fgColor="1C2640")
    hfont = Font(bold=True,color="7C9CFF",name="Calibri")
    for ci,col in enumerate(df.columns,1):
        c = ws.cell(row=1,column=ci,value=str(col))
        c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center")
    for ri,row in enumerate(df.itertuples(index=False),2):
        for ci,val in enumerate(row,1):
            ws.cell(row=ri,column=ci,value=val)
    for ci in range(1,len(df.columns)+1):
        ws.column_dimensions[get_column_letter(ci)].width=22
    wb.save(buf); buf.seek(0)
    return buf.read()

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
DEFAULTS = {"items_df":None,"images":[],"links_df":None,"headings_df":None,
            "tables":[],"meta":{},"ai_info":{},"strategy":"","base_url":""}
for k,v in DEFAULTS.items():
    if k not in st.session_state: st.session_state[k]=v

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="logo" style="font-size:1.4rem;margin-bottom:.2rem;">web2sheet<span style="color:#ec4899;">/</span></div>', unsafe_allow_html=True)
    st.markdown('<div style="font-family:\'DM Mono\',monospace;font-size:.6rem;color:#6d5f8a;letter-spacing:2px;text-transform:uppercase;margin-bottom:.4rem;">URL → Spreadsheet</div>', unsafe_allow_html=True)
    st.markdown('<hr class="divider">', unsafe_allow_html=True)

    # ── ScraperAPI key input ─────────────────────────────────────────────────
    st.markdown("#### 🔑 ScraperAPI Key")
    st.markdown("""<div style="font-size:.75rem;color:#9aa4c7;margin-bottom:.4rem;">
    Unlocks 91mobiles, Flipkart, news sites & more.
    <a href="https://scraperapi.com" target="_blank" style="color:#7c9cff;">Get free key →</a>
    (1,000 req/month free)
    </div>""", unsafe_allow_html=True)

    # Pull from secrets first, allow override
    default_key = ""
    try: default_key = st.secrets.get("SCRAPERAPI_KEY","")
    except Exception: pass

    scraper_key = st.text_input("ScraperAPI Key", value=default_key,
                                 placeholder="Enter key to bypass blocks...",
                                 type="password", label_visibility="collapsed")

    if scraper_key:
        st.success("✅ ScraperAPI active")
    else:
        st.warning("⚠️ No key — only open sites work")

    st.markdown('<hr class="divider">', unsafe_allow_html=True)
    st.markdown("#### ⚙️ Extract")
    use_ai      = st.toggle("🧠 AI Analysis", value=True)
    ai_fallback = st.toggle("🤖 AI Search Fallback", value=True,
                            help="If all fetch methods fail, use Claude web search to extract data")
    ext_imgs    = st.toggle("🖼️ Images", value=True)
    ext_links   = st.toggle("🔗 Links", value=True)
    ext_hdrs    = st.toggle("📝 Headings", value=True)
    ext_tbls    = st.toggle("📊 Tables", value=True)
    trim_ws     = st.toggle("✂️ Trim Whitespace", value=True)
    dedup       = st.toggle("🗑️ De-duplicate", value=False)
    rate_delay  = st.slider("Delay (s)", 0.5, 4.0, 1.0, 0.5)
    max_imgs    = st.slider("Max image previews", 8, 60, 24, 4)

    st.markdown('<hr class="divider">', unsafe_allow_html=True)
    st.caption("PII detection · Rate limited")

# ─────────────────────────────────────────────────────────────────────────────
# HERO
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div style="text-align:center;padding:1.4rem 1rem .5rem;">
  <div class="logo-wrap" data-t="web2sheet">
    <div class="logo">web2sheet</div>
  </div>
  <div style="font-family:'DM Mono',monospace;color:#6d5f8a;font-size:.76rem;margin-top:.4rem;letter-spacing:3px;text-transform:uppercase;">
    // paste a url · get a spreadsheet
  </div>
</div>""", unsafe_allow_html=True)

# ── URL bar ───────────────────────────────────────────────────────────────────
cu, cb = st.columns([5,1])
with cu:
    url_input = st.text_input("URL", placeholder="https://91mobiles.com/... or any page",
                               label_visibility="collapsed")
with cb:
    go = st.button("⚡ Extract", use_container_width=True, type="primary")

# ── No-key notice ─────────────────────────────────────────────────────────────
if not scraper_key and not go:
    st.markdown("""
<div class="key-banner">
  <b style="color:#c4b5fd;">💡 Add a free ScraperAPI key for best results.</b><br>
  <span style="color:#6d5f8a;font-size:.76rem;">
  Without it, only open sites work reliably.<br>
  With it: 91mobiles, Flipkart, SmartPrix & most others work too.<br>
  <a href="https://scraperapi.com" target="_blank" style="color:#a78bfa;">Get free key at scraperapi.com →</a>
  </span>
</div>""", unsafe_allow_html=True)

st.markdown('<hr class="divider">', unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────
if go and url_input:
    for k,v in DEFAULTS.items(): st.session_state[k]=v

    client = None
    try: client = Anthropic()
    except Exception: pass

    with st.status("⚡ Extracting...", expanded=True) as status:

        # ── Fetch ──────────────────────────────────────────────────────────
        st.write("🌐 Fetching page...")
        strategy_labels = {
            "scraperapi":  "✅ Fetched via ScraperAPI",
            "direct":      "✅ Fetched directly",
            "google_cache":"✅ Fetched via Google Cache",
            "wayback":     "✅ Fetched via Wayback Machine",
            "ai_search":   "✅ Data extracted via AI web search (no HTML fetch needed)",
        }
        try:
            html, strategy = smart_fetch(
                url_input, scraper_key, ai_fallback and client is not None,
                client=client, delay=rate_delay
            )
            st.write(strategy_labels.get(strategy, f"✅ Fetched ({strategy})"))
            if strategy == "google_cache":
                st.info("ℹ️ Using Google's cached copy — may be hours/days old.")
            elif strategy == "wayback":
                st.info("ℹ️ Using Wayback Machine archive — may be outdated.")
            elif strategy == "ai_search":
                st.info("ℹ️ Direct fetch blocked — Claude searched the web and extracted data instead.")
            st.session_state.strategy = strategy
        except Exception as e:
            st.error(f"❌ {e}")
            st.markdown("""
**What to do:**
- Add a **free ScraperAPI key** in the sidebar (fixes 90% of blocked sites)
- Try an open site: `books.toscrape.com`, `quotes.toscrape.com`, `en.wikipedia.org`
- Enable **AI Search Fallback** in the sidebar
""")
            status.update(label="❌ Fetch failed", state="error")
            st.stop()

        soup = BeautifulSoup(html, "html.parser")
        st.session_state.base_url = url_input
        st.write(f"📄 {len(html):,} characters parsed")

        # ── AI Analysis ────────────────────────────────────────────────────
        if use_ai and client and strategy != "ai_search":
            st.write("🧠 AI analyzing...")
            try:
                ai_info = ai_analyze(html, client)
                st.session_state.ai_info = ai_info
                st.write(f"✅ {ai_info.get('page_type','?')} — {ai_info.get('summary','')[:90]}")
            except Exception as e:
                st.write(f"⚠️ AI skipped: {e}")

        # ── Structured items ───────────────────────────────────────────────
        st.write("🃏 Extracting items...")
        items = extract_items(soup, url_input)
        if items:
            df = pd.DataFrame(items)
            df = clean_df(df, trim_ws)
            if dedup: df = df.drop_duplicates(subset=["Title"]).reset_index(drop=True)
            pii_cols = [c for c in df.columns if any(k in c.lower() for k in
                        ["email","phone","address","ssn","dob","national"])]
            if pii_cols: st.warning(f"🔒 PII detected: `{', '.join(pii_cols)}`")
            st.session_state.items_df = df
            st.write(f"✅ {len(df)} items · {len(df.columns)} fields")
        else:
            st.write("ℹ️ No repeating items found")

        # ── Images ────────────────────────────────────────────────────────
        if ext_imgs:
            imgs = extract_images(soup, url_input)
            st.session_state.images = imgs
            st.write(f"✅ {len(imgs)} images")

        # ── Tables ────────────────────────────────────────────────────────
        if ext_tbls:
            tbls = extract_tables(soup)
            st.session_state.tables = tbls
            st.write(f"✅ {len(tbls)} tables")

        # ── Links ─────────────────────────────────────────────────────────
        if ext_links:
            lnks = extract_links(soup, url_input)
            st.session_state.links_df = pd.DataFrame(lnks) if lnks else pd.DataFrame()
            st.write(f"✅ {len(lnks)} links")

        # ── Headings ──────────────────────────────────────────────────────
        if ext_hdrs:
            hdgs = extract_headings(soup)
            st.session_state.headings_df = pd.DataFrame(hdgs) if hdgs else pd.DataFrame()
            st.write(f"✅ {len(hdgs)} headings")

        # ── Meta ──────────────────────────────────────────────────────────
        st.session_state.meta = extract_meta(soup)

        status.update(label="✅ Done!", state="complete")

# ─────────────────────────────────────────────────────────────────────────────
# RESULTS
# ─────────────────────────────────────────────────────────────────────────────
has_data = any([st.session_state.items_df is not None,
                st.session_state.images,
                st.session_state.links_df is not None,
                st.session_state.tables])

if has_data:
    ai   = st.session_state.ai_info
    base = st.session_state.base_url
    items_df = st.session_state.items_df
    images   = st.session_state.images
    links_df = st.session_state.links_df
    tables   = st.session_state.tables
    hdgs_df  = st.session_state.headings_df
    meta     = st.session_state.meta

    # AI banner
    if ai and ai.get("page_type"):
        icons = {"e-commerce":"🛒","news":"📰","directory":"📋",
                 "blog":"✍️","jobs":"💼","real-estate":"🏠","other":"🌐"}
        icon = icons.get(ai.get("page_type","other"),"🌐")
        cats_html = " ".join(f'<span class="pill">{c}</span>' for c in ai.get("categories",[])[:10])
        st.markdown(f"""
<div class="ai-banner">
  <span style="font-size:1.4rem;">{icon}</span>
  <span style="font-family:'Syne',sans-serif;font-weight:700;margin-left:10px;color:#a78bfa;">
    {ai.get("page_type","").replace("-"," ").title()}
  </span>
  <span style="color:#6d5f8a;font-size:.78rem;margin-left:10px;">{ai.get("summary","")}</span>
  <div style="margin-top:.5rem;">{cats_html}</div>
</div>""", unsafe_allow_html=True)

    # Strategy badge
    strat = st.session_state.strategy
    strat_color = {"scraperapi":"#22d3ee","direct":"#28c840",
                   "google_cache":"#f472b6","wayback":"#f472b6","ai_search":"#a78bfa"}
    if strat:
        sc = strat_color.get(strat,"#6d5f8a")
        st.markdown(f'<span class="strat-badge" style="border-color:{sc};color:{sc};">⬡ Source: {strat.replace("_"," ").upper()}</span>',
                    unsafe_allow_html=True)

    # Stats
    c1,c2,c3,c4,c5 = st.columns(5)
    c1.markdown(f'<div class="sbox"><div class="snum">{len(items_df) if items_df is not None else 0}</div><div class="slbl">Items</div></div>',unsafe_allow_html=True)
    c2.markdown(f'<div class="sbox"><div class="snum">{len(images)}</div><div class="slbl">Images</div></div>',unsafe_allow_html=True)
    c3.markdown(f'<div class="sbox"><div class="snum">{len(links_df) if links_df is not None else 0}</div><div class="slbl">Links</div></div>',unsafe_allow_html=True)
    c4.markdown(f'<div class="sbox"><div class="snum">{len(tables)}</div><div class="slbl">Tables</div></div>',unsafe_allow_html=True)
    c5.markdown(f'<div class="sbox"><div class="snum">{len(hdgs_df) if hdgs_df is not None else 0}</div><div class="slbl">Headings</div></div>',unsafe_allow_html=True)
    st.markdown('<hr class="divider">', unsafe_allow_html=True)

    # ── TABS ──────────────────────────────────────────────────────────────────
    tabs = st.tabs(["🃏 Items","🖼️ Images","📊 Tables","🔗 Links","📝 Headings","🏷️ Meta","📥 Export"])

    # ── Items ──────────────────────────────────────────────────────────────────
    with tabs[0]:
        if items_df is not None and not items_df.empty:
            df = items_df.copy()
            view = st.radio("View",["🃏 Cards","📋 Table","🔍 Inspector"],horizontal=True,label_visibility="collapsed")
            sc1,sc2 = st.columns([3,2])
            with sc1:
                srch = st.text_input("Search","",placeholder="Filter...",label_visibility="collapsed")
            with sc2:
                cats = ["All"]+sorted([x for x in df.get("Category",pd.Series()).dropna().unique() if x])
                sel_cat = st.selectbox("Cat",cats,label_visibility="collapsed") if "Category" in df.columns else "All"
            if srch:
                m = df.apply(lambda r: r.astype(str).str.contains(srch,case=False).any(),axis=1)
                df = df[m]
            if sel_cat != "All" and "Category" in df.columns:
                df = df[df["Category"]==sel_cat]
            st.caption(f"{len(df)} of {len(items_df)} items")

            if view == "🃏 Cards":
                for i in range(0, len(df), 3):
                    chunk = df.iloc[i:i+3]
                    cols = st.columns(3)
                    for ci,(_, item) in enumerate(chunk.iterrows()):
                        with cols[ci]:
                            img_url = str(item.get("Image URL",""))
                            title   = str(item.get("Title","—"))[:90]
                            price   = str(item.get("Price",""))
                            rating  = str(item.get("Rating",""))
                            desc    = str(item.get("Description",""))[:160]
                            url_    = str(item.get("URL",""))
                            cat     = str(item.get("Category",""))
                            brand   = str(item.get("Brand",""))
                            avail   = str(item.get("Availability",""))

                            img_h = (f'<img src="{img_url}" style="width:100%;height:165px;object-fit:cover;border-radius:12px 12px 0 0;display:block;" onerror="this.style.display=\'none\'">'
                                     if img_url and img_url!="nan" else
                                     '<div style="width:100%;height:70px;background:rgba(124,156,255,0.07);border-radius:12px 12px 0 0;display:flex;align-items:center;justify-content:center;font-size:2rem;">📦</div>')
                            parts = []
                            if cat and cat!="nan":   parts.append(f'<span class="pill">{cat}</span>')
                            parts.append(f'<div class="ptitle">{title}</div>')
                            if brand and brand!="nan": parts.append(f'<div style="font-size:.7rem;color:#9aa4c7;">by {brand}</div>')
                            if price and price!="nan": parts.append(f'<div class="pprice">{price}</div>')
                            if rating and rating!="nan": parts.append(f'<div class="prating">⭐ {rating}</div>')
                            if avail and avail!="nan": parts.append(f'<div style="font-size:.7rem;color:#28c840;">{avail}</div>')
                            if desc and desc!="nan": parts.append(f'<div class="pdesc">{desc}</div>')
                            if url_ and url_!="nan": parts.append(f'<a href="{url_}" target="_blank" style="color:#7c9cff;font-size:.74rem;">🔗 Visit →</a>')
                            body = "".join(parts)
                            st.markdown(f'<div class="pcard">{img_h}<div class="pbody">{body}</div></div>',unsafe_allow_html=True)

            elif view == "📋 Table":
                all_c = list(df.columns)
                show  = st.multiselect("Columns",all_c,default=all_c[:9],key="vis")
                cfg   = {c: st.column_config.LinkColumn(c) for c in (show or all_c) if "url" in c.lower()}
                st.dataframe(df[show] if show else df, use_container_width=True, height=520, column_config=cfg)

            else:  # Inspector
                idx = st.slider("Item",1,max(len(df),1),1)-1
                item = df.iloc[idx]
                l,r = st.columns([1,2])
                with l:
                    iu = str(item.get("Image URL",""))
                    if iu and iu!="nan":
                        st.markdown(f'<img src="{iu}" style="width:100%;border-radius:12px;border:1px solid rgba(255,255,255,0.08);" onerror="this.style.display=\'none\'">',unsafe_allow_html=True)
                    else:
                        st.markdown('<div style="width:100%;height:180px;background:rgba(124,156,255,0.07);border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:3rem;">📦</div>',unsafe_allow_html=True)
                with r:
                    rows_h = ""
                    for col,val in item.items():
                        vs = str(val)
                        if not vs or vs in ["nan","","None"] or col=="Image URL": continue
                        if col=="URL": vs=f'<a href="{vs}" target="_blank" style="color:#7c9cff;">{vs[:70]}</a>'
                        rows_h += f"<tr><td>{col}</td><td>{vs}</td></tr>"
                    st.markdown(f'<table class="dtable"><tbody>{rows_h}</tbody></table>',unsafe_allow_html=True)
        else:
            st.info("No structured items found on this page.")

    # ── Images ─────────────────────────────────────────────────────────────────
    with tabs[1]:
        imgs = st.session_state.images
        if imgs:
            st.markdown(f"### {len(imgs)} Images")
            fc1,fc2 = st.columns([3,2])
            with fc1: ifilt = st.text_input("Filter","",placeholder="jpg, logo, product...",label_visibility="collapsed")
            with fc2: ncols = st.select_slider("Cols",[2,3,4,5,6],value=4)
            disp = [i for i in imgs if ifilt.lower() in i["Image URL"].lower()][:max_imgs] if ifilt else imgs[:max_imgs]
            st.caption(f"{len(disp)} of {len(imgs)}")
            for rs in range(0,len(disp),ncols):
                row_imgs = disp[rs:rs+ncols]
                c_ = st.columns(ncols)
                for ci,img in enumerate(row_imgs):
                    with c_[ci]:
                        u = img["Image URL"]
                        alt = img.get("Alt Text","")
                        st.markdown(f'<div style="margin-bottom:.7rem;"><img src="{u}" style="width:100%;height:125px;object-fit:cover;border-radius:8px;border:1px solid rgba(255,255,255,0.06);" onerror="this.style.display=\'none\'"><div style="font-size:.63rem;color:#9aa4c7;margin-top:2px;overflow:hidden;white-space:nowrap;text-overflow:ellipsis;">{alt or u.split("/")[-1][:30]}</div></div>',unsafe_allow_html=True)
            with st.expander(f"📋 All {len(imgs)} URLs"):
                idf = pd.DataFrame(imgs)
                st.dataframe(idf,use_container_width=True,height=250,
                             column_config={"Image URL":st.column_config.LinkColumn("Image URL")})
            idf2 = pd.DataFrame(imgs)
            c1_,c2_ = st.columns(2)
            c1_.download_button("⬇️ CSV",idf2.to_csv(index=False).encode(),"w2s_images.csv","text/csv")
            c2_.download_button("⬇️ JSON",idf2.to_json(orient="records",indent=2).encode(),"w2s_images.json","application/json")
        else:
            st.info("No images found.")

    # ── Tables ─────────────────────────────────────────────────────────────────
    with tabs[2]:
        tbls = st.session_state.tables
        if tbls:
            st.markdown(f"### {len(tbls)} Tables")
            for i,t in enumerate(tbls):
                with st.expander(f"Table {i+1} — {len(t)} rows × {len(t.columns)} cols",expanded=(i==0)):
                    st.dataframe(t,use_container_width=True,height=280)
                    tc1,tc2,tc3 = st.columns(3)
                    tc1.download_button("⬇️ CSV",t.to_csv(index=False).encode(),f"w2s_table{i+1}.csv","text/csv",key=f"tc{i}")
                    tc2.download_button("⬇️ JSON",t.to_json(orient="records",indent=2).encode(),f"w2s_table{i+1}.json","application/json",key=f"tj{i}")
                    try: tc3.download_button("⬇️ Excel",to_excel(t),f"w2s_table{i+1}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key=f"tx{i}")
                    except Exception: pass
        else:
            st.info("No HTML tables found.")

    # ── Links ──────────────────────────────────────────────────────────────────
    with tabs[3]:
        ldf = st.session_state.links_df
        if ldf is not None and not ldf.empty:
            st.markdown(f"### {len(ldf)} Links")
            ls = st.text_input("Filter","",placeholder="Search...",label_visibility="collapsed")
            dl = ldf[ldf.apply(lambda r:r.astype(str).str.contains(ls,case=False).any(),axis=1)] if ls else ldf
            st.dataframe(dl,use_container_width=True,height=400,
                         column_config={"URL":st.column_config.LinkColumn("URL")})
            lc1,lc2 = st.columns(2)
            lc1.download_button("⬇️ CSV",dl.to_csv(index=False).encode(),"w2s_links.csv","text/csv")
            try: lc2.download_button("⬇️ Excel",to_excel(dl),"w2s_links.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception: pass
            with st.expander("🌍 Domains"):
                dc = dl["Domain"].value_counts().reset_index(); dc.columns=["Domain","Links"]
                st.dataframe(dc,use_container_width=True,height=180)
        else:
            st.info("No links found.")

    # ── Headings ───────────────────────────────────────────────────────────────
    with tabs[4]:
        hdf = st.session_state.headings_df
        if hdf is not None and not hdf.empty:
            st.markdown(f"### {len(hdf)} Headings")
            lcolors={"H1":"#c4b5fd","H2":"#a78bfa","H3":"#f472b6","H4":"#6d5f8a","H5":"#6d5f8a","H6":"#6d5f8a"}
            for lv in ["H1","H2","H3","H4","H5","H6"]:
                sub = hdf[hdf["Level"]==lv]
                if sub.empty: continue
                st.markdown(f'<div class="cheader"><span style="color:{lcolors.get(lv,"#9aa4c7")};font-family:Poppins,sans-serif;font-weight:700;">{lv}</span><span class="ccount">{len(sub)}</span></div>',unsafe_allow_html=True)
                for _,r in sub.iterrows():
                    sz={"H1":"1rem","H2":".9rem","H3":".84rem"}.get(lv,".78rem")
                    st.markdown(f'<div style="padding:.3rem .8rem;margin-bottom:.2rem;background:rgba(18,24,38,0.5);border-radius:6px;font-size:{sz};">{r["Text"]}</div>',unsafe_allow_html=True)
            st.download_button("⬇️ CSV",hdf.to_csv(index=False).encode(),"w2s_headings.csv","text/csv")
        else:
            st.info("No headings found.")

    # ── Meta ───────────────────────────────────────────────────────────────────
    with tabs[5]:
        meta = st.session_state.meta
        if meta:
            st.markdown("### Page Metadata")
            rows_h="".join(f"<tr><td>{k}</td><td>{str(v)[:350]}</td></tr>" for k,v in meta.items() if v)
            st.markdown(f'<table class="dtable"><thead><tr><th>Field</th><th>Value</th></tr></thead><tbody>{rows_h}</tbody></table>',unsafe_allow_html=True)
            mdf=pd.DataFrame(list(meta.items()),columns=["Field","Value"])
            st.download_button("⬇️ CSV",mdf.to_csv(index=False).encode(),"w2s_meta.csv","text/csv")
        else:
            st.info("No metadata found.")

    # ── Export All ─────────────────────────────────────────────────────────────
    with tabs[6]:
        st.markdown("### Export Everything")
        def exp_block(label, df_e, pfx):
            if df_e is None or df_e.empty: return
            st.markdown(f"**{label}** — {len(df_e):,} rows")
            ec1,ec2,ec3=st.columns(3)
            ec1.download_button("⬇️ CSV",df_e.to_csv(index=False).encode(),f"w2s_{pfx}.csv","text/csv",key=f"ec{pfx}")
            ec2.download_button("⬇️ JSON",df_e.to_json(orient="records",indent=2).encode(),f"w2s_{pfx}.json","application/json",key=f"ej{pfx}")
            try: ec3.download_button("⬇️ Excel",to_excel(df_e),f"w2s_{pfx}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key=f"ex{pfx}")
            except Exception: pass
            st.markdown('<hr class="divider">',unsafe_allow_html=True)

        exp_block("🃏 Items", st.session_state.items_df, "items")
        if st.session_state.images:
            exp_block("🖼️ Images", pd.DataFrame(st.session_state.images), "images")
        exp_block("🔗 Links", st.session_state.links_df, "links")
        exp_block("📝 Headings", st.session_state.headings_df, "headings")
        for i,t in enumerate(st.session_state.tables or []):
            exp_block(f"📊 Table {i+1}", t, f"table{i+1}")
        if st.session_state.meta:
            exp_block("🏷️ Meta", pd.DataFrame(list(st.session_state.meta.items()),columns=["Field","Value"]), "meta")

# ─────────────────────────────────────────────────────────────────────────────
# EMPTY STATE
# ─────────────────────────────────────────────────────────────────────────────
elif not go:
    c1,c2,c3,c4 = st.columns(4)
    tiles = [
        ("🖼️","Images","Every image URL, alt text, dimensions — gallery + CSV"),
        ("🃏","Products","Title · Price · Rating · Image · Brand · SKU · URL"),
        ("📊","Tables","All HTML tables → clean DataFrames"),
        ("🔗","Links & Meta","All links, headings, page metadata"),
    ]
    for col,(icon,title,desc) in zip([c1,c2,c3,c4],tiles):
        with col:
            st.markdown(f'<div class="glass" style="text-align:center;"><div style="font-size:2rem;">{icon}</div><div style="font-family:Poppins,sans-serif;font-weight:700;margin:.4rem 0 .3rem;">{title}</div><div style="color:#9aa4c7;font-size:.78rem;">{desc}</div></div>',unsafe_allow_html=True)

    st.markdown('<hr class="divider">', unsafe_allow_html=True)
    st.markdown("#### 💡 Try These")
    ex = [
        ("📚","books.toscrape.com","https://books.toscrape.com/"),
        ("💬","quotes.toscrape.com","https://quotes.toscrape.com/"),
        ("📰","HackerNews","https://news.ycombinator.com/"),
        ("🌍","Wikipedia tables","https://en.wikipedia.org/wiki/List_of_countries_by_GDP_(nominal)"),
        ("📱","91mobiles*","https://www.91mobiles.com/mobile-phones-under-20000"),
        ("🛒","Flipkart*","https://www.flipkart.com/mobiles"),
    ]
    ec1,ec2,ec3 = st.columns(3)
    for i,(icon,name,url_) in enumerate(ex):
        with [ec1,ec2,ec3][i%3]:
            note = " <span style='font-size:.62rem;color:#f472b6;letter-spacing:1px;'>*needs ScraperAPI</span>" if "*" in name else ""
            name_clean = name.replace("*","")
            st.markdown(f'<div class="glass" style="padding:.75rem 1rem;border-left:2px solid rgba(139,92,246,0.4);"><b style="color:#c4b5fd;">{icon} {name_clean}</b>{note}<br><span style="font-size:.68rem;color:#6d5f8a;">{url_}</span></div>',unsafe_allow_html=True)

# ── FOOTER ────────────────────────────────────────────────────────────────────
st.markdown("""
<hr class="divider" style="margin-top:2rem;">
<div style="display:flex;align-items:center;justify-content:space-between;padding:.8rem 0 1.6rem;flex-wrap:wrap;gap:12px;">
  <div style="font-family:'Space Mono',monospace;font-weight:700;font-size:.85rem;
    background:linear-gradient(135deg,#c4b5fd,#f472b6);
    -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;">
    web2sheet<span style="-webkit-text-fill-color:#ec4899;">/</span>
  </div>
  <div style="display:flex;gap:24px;">
    <a href="https://scraperapi.com" target="_blank"
      style="font-family:'DM Mono',monospace;font-size:.58rem;letter-spacing:2px;
             text-transform:uppercase;color:#6d5f8a;text-decoration:none;
             transition:color .2s;">ScraperAPI</a>
    <a href="https://docs.anthropic.com" target="_blank"
      style="font-family:'DM Mono',monospace;font-size:.58rem;letter-spacing:2px;
             text-transform:uppercase;color:#6d5f8a;text-decoration:none;">Anthropic</a>
    <a href="https://streamlit.io" target="_blank"
      style="font-family:'DM Mono',monospace;font-size:.58rem;letter-spacing:2px;
             text-transform:uppercase;color:#6d5f8a;text-decoration:none;">Streamlit</a>
  </div>
  <div style="font-family:'DM Mono',monospace;font-size:.54rem;color:#6d5f8a;opacity:.5;letter-spacing:1px;">
    © 2025 WEB2SHEET · PII DETECTION · RATE LIMITED
  </div>
</div>
""", unsafe_allow_html=True)
